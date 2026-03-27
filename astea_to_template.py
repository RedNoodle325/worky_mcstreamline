#!/usr/bin/env python3
"""
Convert an Astea equipment export CSV into the Unit_Import_Template.xlsx format.

Usage:
    python astea_to_template.py <input_csv> [output_xlsx]

If output_xlsx is omitted, saves alongside input as <same_name>_template.xlsx.
"""

import csv
import sys
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

TEMPLATE_PATH = "/Users/zakklinedinst/Work/worky_mcstreamline/Unit_Import_Template.xlsx"

FILL_ODD = PatternFill("solid", fgColor="F0F4FF")   # light blue for odd data rows
FILL_EVEN = PatternFill("solid", fgColor="FFFFFF")  # white for even data rows

UNIT_TYPE_MAP = {
    "COND": "ACCU",
    "EVAP": "CRAC",
}


def parse_install_date(raw: str) -> str:
    raw = raw.strip()
    if not raw:
        return ""
    try:
        dt = datetime.strptime(raw, "%Y-%m-%d %I:%M:%S %p")
        return dt.strftime("%m/%d/%Y")
    except ValueError:
        return ""


def bpart_suffix_to_unit_type(bpart_id: str) -> str:
    suffix = bpart_id.strip().rsplit("-", 1)[-1].upper()
    return UNIT_TYPE_MAP.get(suffix, suffix)


def strip_location_prefix(pclass2_descr: str) -> str:
    val = pclass2_descr.strip()
    if ": " in val:
        return val.split(": ", 1)[1].strip()
    return val


def map_manufacturer(def_node_id: str) -> str:
    val = def_node_id.strip()
    return "Munters" if "MUNTERS" in val.upper() else val


def parse_row(row: list) -> dict:
    # Data rows have 22 cols due to embedded comma in descr and bpart_descr
    bpart_id       = row[0].strip()
    company_id     = row[1].strip()
    descr          = (row[2] + "," + row[3]).strip()
    install_date   = row[4].strip()
    serial_no      = row[5].strip()
    def_node_id    = row[8].strip()
    pclass2_descr  = row[19].strip()
    company_descr  = row[15].strip()

    return {
        "site_name":        "",
        "unit_type":        bpart_suffix_to_unit_type(bpart_id),
        "asset_tag":        bpart_id,
        "manufacturer":     map_manufacturer(def_node_id),
        "model":            descr,
        "serial_number":    serial_no,
        "job_number":       company_id,
        "line_number":      "",
        "location_in_site": strip_location_prefix(pclass2_descr),
        "description":      company_descr,
        "install_date":     parse_install_date(install_date),
        "warranty_start":   "",
        "warranty_end":     "",
        "notes":            "",
    }


def main():
    if len(sys.argv) < 2:
        print("Usage: python astea_to_template.py <input_csv> [output_xlsx]")
        sys.exit(1)

    input_csv = sys.argv[1]
    if len(sys.argv) >= 3:
        output_xlsx = sys.argv[2]
    else:
        base = os.path.splitext(input_csv)[0]
        output_xlsx = base + "_template.xlsx"

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Units"]

    # Collect data rows from CSV (skip header)
    data_rows = []
    with open(input_csv, newline="", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        next(reader)  # skip header
        for row in reader:
            if len(row) < 20:
                continue
            data_rows.append(parse_row(row))

    # Write starting at row 3 (row 1 = header, row 2 = example yellow row)
    for i, record in enumerate(data_rows):
        excel_row = i + 3  # 1-indexed, starting at row 3
        fill = FILL_ODD if i % 2 == 0 else FILL_EVEN  # i=0 → row 3 (odd)

        values = [
            record["site_name"],
            record["unit_type"],
            record["asset_tag"],
            record["manufacturer"],
            record["model"],
            record["serial_number"],
            record["job_number"],
            record["line_number"],
            record["location_in_site"],
            record["description"],
            record["install_date"],
            record["warranty_start"],
            record["warranty_end"],
            record["notes"],
        ]

        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=excel_row, column=col_idx, value=val if val != "" else None)
            cell.fill = fill

    wb.save(output_xlsx)
    print(f"Rows written: {len(data_rows)}")
    print(f"Output file:  {output_xlsx}")


if __name__ == "__main__":
    main()
